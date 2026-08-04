from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, make_response
import os
from dotenv import load_dotenv
load_dotenv()
import pymysql
pymysql.install_as_MySQLdb()
from flask_mysqldb import MySQL
from werkzeug.middleware.proxy_fix import ProxyFix
import time
from collections import Counter
from functools import wraps
from urllib.parse import urlparse
from flask_mail import Mail, Message
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf import CSRFProtect
from flask_wtf.csrf import CSRFError
import secrets
from werkzeug.security import generate_password_hash, check_password_hash
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from io import BytesIO
import csv
import io
import gzip
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── APP CONFIG ───────────────────────────────────────────────────────────────
app = Flask(__name__, static_folder='static', static_url_path='/static')
# Render (and most PaaS) terminate TLS at the edge and forward internally over
# plain HTTP -- without this, request.remote_addr is always the proxy's IP
# (breaking per-client rate limiting) and url_for(_external=True) generates
# http:// links instead of https://.
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
app.config['MAX_CONTENT_LENGTH']         = 16 * 1024 * 1024  # 16MB request body cap
app.config['SEND_FILE_MAX_AGE_DEFAULT']  = 31536000
app.config['PERMANENT_SESSION_LIFETIME'] = 1800
app.config['SESSION_COOKIE_HTTPONLY']    = True
app.config['SESSION_COOKIE_SECURE']      = True
app.config['SESSION_COOKIE_SAMESITE']   = 'Lax'
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24).hex())

csrf = CSRFProtect(app)

# ── DATABASE ─────────────────────────────────────────────────────────────────
app.config['MYSQL_HOST']     = os.environ.get('MYSQL_HOST')
app.config['MYSQL_USER']     = os.environ.get('MYSQL_USER')
app.config['MYSQL_PASSWORD'] = os.environ.get('MYSQL_PASSWORD')
app.config['MYSQL_PORT']     = int(os.environ.get('MYSQL_PORT', 3306))
app.config['MYSQL_DB']       = os.environ.get('MYSQL_DB')
mysql = MySQL(app)

# ── MAIL ─────────────────────────────────────────────────────────────────────
app.config['MAIL_SERVER']   = 'smtp.gmail.com'
app.config['MAIL_PORT']     = 587
app.config['MAIL_USE_TLS']  = True
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
mail = Mail(app)

# ── RATE LIMITER ─────────────────────────────────────────────────────────────
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["300 per day", "60 per hour"],
    storage_uri="memory://"
)

# ── RESET TOKENS (with expiry) ───────────────────────────────────────────────
reset_tokens = {}  # {token: {'email': str, 'expires': float}}


# ── AFTER REQUEST — Security headers + Gzip ──────────────────────────────────
@app.after_request
def after_request(response):
    # Security headers
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options']         = 'SAMEORIGIN'
    response.headers['X-XSS-Protection']        = '1; mode=block'
    response.headers['Referrer-Policy']         = 'strict-origin-when-cross-origin'
    if request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    # Cache
    if request.endpoint == 'static':
        response.headers['Cache-Control'] = 'public, max-age=31536000'
    else:
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    # Gzip
    if (200 <= response.status_code < 300
            and 'Content-Encoding' not in response.headers
            and len(response.get_data()) >= 500
            and 'gzip' in request.headers.get('Accept-Encoding', '')):
        try:
            buf = io.BytesIO()
            with gzip.GzipFile(mode='wb', fileobj=buf) as f:
                f.write(response.get_data())
            response.set_data(buf.getvalue())
            response.headers['Content-Encoding'] = 'gzip'
            response.headers['Content-Length']   = len(response.get_data())
        except Exception:
            pass
    return response


# ── ERROR HANDLERS ────────────────────────────────────────────────────────────
def safe_redirect_back(default_endpoint):
    """Redirect to the referring page only if it's on this site.

    request.referrer is attacker-controlled (it's just a header the client
    sends), so blindly trusting it -- especially in the CSRF error handler,
    which fires exactly when a cross-site request was blocked -- would bounce
    the user straight to whatever site sent the forged request.
    """
    ref = request.referrer
    if ref and urlparse(ref).netloc == urlparse(request.host_url).netloc:
        return redirect(ref)
    return redirect(url_for(default_endpoint))

@app.errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def server_error(e):
    return render_template('500.html'), 500

@app.errorhandler(429)
def rate_limited(e):
    flash('Too many attempts! Please wait a minute and try again.', 'danger')
    return redirect(url_for('login'))

@app.errorhandler(413)
def too_large(e):
    flash('That file is too large. Please keep uploads under 16MB.', 'danger')
    return safe_redirect_back('dashboard')

@app.errorhandler(CSRFError)
def csrf_error(e):
    flash('Your session expired. Please try that again.', 'danger')
    return safe_redirect_back('login')


# ── AUTH DECORATORS ──────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Access denied!', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


# ── INDEX ─────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if 'logged_in' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


# ── LOGIN ─────────────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute", methods=["POST"])
def login():
    if request.method == 'POST':
        email    = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        if not email or not password:
            flash('Please fill in all fields.', 'danger')
            return render_template('login.html')
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT student_id, name, role, password FROM students WHERE email=%s",
            (email,)
        )
        user = cur.fetchone()
        cur.close()
        if user and check_password_hash(user[3], password):
            session.permanent    = True
            session['logged_in'] = True
            session['user_id']   = user[0]
            session['user_name'] = user[1]
            session['role']      = user[2]
            return redirect(url_for('dashboard'))
        flash('Invalid email or password.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── FORGOT PASSWORD ───────────────────────────────────────────────────────────
@app.route('/forgot_password', methods=['GET', 'POST'])
@limiter.limit("3 per minute", methods=["POST"])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        cur   = mysql.connection.cursor()
        cur.execute("SELECT student_id, name FROM students WHERE email=%s", (email,))
        user  = cur.fetchone()
        cur.close()
        if user:
            token = secrets.token_urlsafe(32)
            reset_tokens[token] = {
                'email':   email,
                'expires': time.time() + 3600  # 1 hour
            }
            reset_link = url_for('reset_password', token=token, _external=True)
            try:
                msg = Message(
                    'Password Reset – Placement Analytics',
                    sender=os.environ.get('MAIL_USERNAME'),
                    recipients=[email]
                )
                msg.body = f'''Hello {user[1]},

Click the link below to reset your password (valid 1 hour):
{reset_link}

This link works only once.
If you did not request this, ignore this email.

Regards, Placement Analytics Team'''
                mail.send(msg)
            except Exception:
                pass
        # Don't reveal whether email exists (security best practice)
        flash('If that email exists, a reset link has been sent.', 'success')
    return render_template('forgot_password.html')


@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if token not in reset_tokens:
        flash('Invalid or expired link!', 'danger')
        return redirect(url_for('login'))
    token_data = reset_tokens[token]
    if time.time() > token_data['expires']:
        del reset_tokens[token]
        flash('Reset link expired! Please request a new one.', 'danger')
        return redirect(url_for('forgot_password'))
    if request.method == 'POST':
        new_password = request.form.get('password', '')
        if len(new_password) < 6:
            flash('Password must be at least 6 characters.', 'danger')
            return render_template('reset_password.html', token=token)
        email  = token_data['email']
        hashed = generate_password_hash(new_password)
        cur    = mysql.connection.cursor()
        cur.execute("UPDATE students SET password=%s WHERE email=%s", (hashed, email))
        mysql.connection.commit()
        cur.close()
        del reset_tokens[token]
        flash('Password reset successful! Please login.', 'success')
        return redirect(url_for('login'))
    return render_template('reset_password.html', token=token)


# ── DASHBOARD ─────────────────────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    cur = mysql.connection.cursor()
    # Single optimized query
    cur.execute("""
        SELECT
            (SELECT COUNT(*) FROM students)   AS ts,
            (SELECT COUNT(*) FROM companies)  AS tc,
            (SELECT COUNT(*) FROM placements) AS tp,
            (SELECT AVG(c.package) FROM placements p JOIN companies c ON p.company_id=c.company_id) AS avg_p,
            (SELECT MAX(c.package) FROM placements p JOIN companies c ON p.company_id=c.company_id) AS max_p
    """)
    s = cur.fetchone()
    total_students   = s[0] or 0
    total_companies  = s[1] or 0
    total_placements = s[2] or 0
    avg_package      = s[3] or 0
    max_package      = s[4] or 0

    cur.execute("""
        SELECT s.name, c.company_name, c.package, p.year, p.status
        FROM placements p
        JOIN students s ON p.student_id=s.student_id
        JOIN companies c ON p.company_id=c.company_id
        ORDER BY p.placement_id DESC LIMIT 5
    """)
    recent = cur.fetchall()

    notifications = []
    if session.get('role') == 'admin':
        cur.execute("""
            SELECT s.name, c.company_name, c.package, p.status
            FROM placements p
            JOIN students s ON p.student_id=s.student_id
            JOIN companies c ON p.company_id=c.company_id
            ORDER BY p.placement_id DESC LIMIT 3
        """)
        for p in cur.fetchall():
            notifications.append({'icon':'🎉', 'message':f'{p[0]} placed at {p[1]} — {p[2]} LPA', 'type':'success'})
        cur.execute("SELECT COUNT(*) FROM students WHERE role='student'")
        ns = cur.fetchone()[0]
        if ns:
            notifications.append({'icon':'👨‍🎓', 'message':f'Total {ns} student(s) registered', 'type':'info'})

    if session.get('role') == 'student':
        cur.execute("""
            SELECT c.company_name, c.package, p.status
            FROM placements p JOIN companies c ON p.company_id=c.company_id
            WHERE p.student_id=%s ORDER BY p.placement_id DESC LIMIT 1
        """, (session['user_id'],))
        mp = cur.fetchone()
        if mp:
            notifications.append({'icon':'🎉', 'message':f'You are placed at {mp[0]} — {mp[1]} LPA!', 'type':'success'})
        else:
            notifications.append({'icon':'💡', 'message':'Check ML Predictor to know your placement chances!', 'type':'info'})

    cur.close()
    return render_template('dashboard.html',
        total_students=total_students, total_companies=total_companies,
        total_placements=total_placements,
        avg_package=round(avg_package, 2), max_package=round(max_package, 2),
        recent=recent, notifications=notifications, user_name=session['user_name'])


# ── STUDENTS ──────────────────────────────────────────────────────────────────
@app.route('/students')
@login_required
def students():
    page     = request.args.get('page', 1, type=int)
    per_page = 20
    offset   = (page - 1) * per_page
    cur      = mysql.connection.cursor()
    cur.execute("SELECT COUNT(*) FROM students")
    total = cur.fetchone()[0]
    cur.execute("""
        SELECT student_id, name, email, branch, cgpa, skills
        FROM students ORDER BY student_id DESC LIMIT %s OFFSET %s
    """, (per_page, offset))
    all_students = cur.fetchall()
    cur.close()
    total_pages = (total + per_page - 1) // per_page
    return render_template('students.html',
        students=all_students, user_name=session['user_name'],
        page=page, total_pages=total_pages, total=total)


@app.route('/add_student', methods=['GET', 'POST'])
@admin_required
def add_student():
    if request.method == 'POST':
        try:
            cgpa = float(request.form['cgpa'])
        except ValueError:
            flash('CGPA must be a number.', 'danger')
            return render_template('add_student.html', user_name=session['user_name'])
        if len(request.form.get('password', '')) < 6:
            flash('Password must be at least 6 characters.', 'danger')
            return render_template('add_student.html', user_name=session['user_name'])
        password = generate_password_hash(request.form['password'])
        cur = mysql.connection.cursor()
        try:
            cur.execute(
                "INSERT INTO students(name,email,branch,cgpa,skills,password) VALUES(%s,%s,%s,%s,%s,%s)",
                (request.form['name'], request.form['email'], request.form['branch'],
                 cgpa, request.form['skills'], password)
            )
            mysql.connection.commit()
            flash('Student added successfully!', 'success')
            return redirect(url_for('students'))
        except pymysql.err.IntegrityError:
            mysql.connection.rollback()
            flash('That email is already registered to another student.', 'danger')
            return render_template('add_student.html', user_name=session['user_name'])
        finally:
            cur.close()
    return render_template('add_student.html', user_name=session['user_name'])


@app.route('/edit_student/<int:student_id>', methods=['GET', 'POST'])
@admin_required
def edit_student(student_id):
    cur = mysql.connection.cursor()
    if request.method == 'POST':
        try:
            cgpa = float(request.form['cgpa'])
        except ValueError:
            cur.close()
            flash('CGPA must be a number.', 'danger')
            return redirect(url_for('edit_student', student_id=student_id))
        try:
            cur.execute("SELECT 1 FROM students WHERE student_id=%s", (student_id,))
            if not cur.fetchone():
                flash('Student not found — they may have already been deleted.', 'danger')
                return redirect(url_for('students'))
            cur.execute("""
                UPDATE students SET name=%s, email=%s, branch=%s, cgpa=%s, skills=%s
                WHERE student_id=%s
            """, (request.form['name'], request.form['email'], request.form['branch'],
                  cgpa, request.form['skills'], student_id))
            mysql.connection.commit()
            flash('Student updated successfully!', 'success')
            return redirect(url_for('students'))
        except pymysql.err.IntegrityError:
            mysql.connection.rollback()
            flash('That email is already registered to another student.', 'danger')
            return redirect(url_for('edit_student', student_id=student_id))
        finally:
            cur.close()
    cur.execute("SELECT * FROM students WHERE student_id=%s", (student_id,))
    student = cur.fetchone()
    cur.close()
    if not student:
        flash('Student not found.', 'danger')
        return redirect(url_for('students'))
    return render_template('edit_student.html', student=student, user_name=session['user_name'])


@app.route('/delete_student/<int:student_id>', methods=['POST'])
@admin_required
def delete_student(student_id):
    cur = mysql.connection.cursor()
    try:
        cur.execute("DELETE FROM students WHERE student_id=%s", (student_id,))
        mysql.connection.commit()
        flash('Student deleted successfully!', 'success')
    except pymysql.err.IntegrityError:
        mysql.connection.rollback()
        flash('Cannot delete this student — they have placement records. Delete those placements first.', 'danger')
    finally:
        cur.close()
    return redirect(url_for('students'))


@app.route('/upload_csv', methods=['GET', 'POST'])
@admin_required
def upload_csv():
    if request.method == 'POST':
        if 'csv_file' not in request.files:
            flash('No file selected!', 'danger')
            return redirect(url_for('upload_csv'))
        file = request.files['csv_file']
        if file.filename == '' or not file.filename.endswith('.csv'):
            flash('Only CSV files allowed!', 'danger')
            return redirect(url_for('upload_csv'))
        try:
            stream  = io.StringIO(file.stream.read().decode('utf-8'))
            rows    = list(csv.DictReader(stream))
            if not rows:
                flash('CSV file is empty!', 'danger')
                return redirect(url_for('upload_csv'))
            required = ['name', 'email', 'branch', 'cgpa', 'skills', 'password']
            missing  = [c for c in required if c not in rows[0].keys()]
            if missing:
                flash(f'Missing columns: {", ".join(missing)}', 'danger')
                return redirect(url_for('upload_csv'))
            cur = mysql.connection.cursor()
            success = errors = 0
            for row in rows:
                try:
                    cur.execute("""
                        INSERT INTO students(name,email,branch,cgpa,skills,password,role)
                        VALUES(%s,%s,%s,%s,%s,%s,'student')
                    """, (str(row['name']), str(row['email']), str(row['branch']),
                          float(row['cgpa']), str(row['skills']),
                          generate_password_hash(str(row['password']))))
                    mysql.connection.commit()
                    success += 1
                except Exception:
                    errors += 1
            cur.close()
            flash(f'✅ {success} students added! ❌ {errors} errors skipped.', 'success')
            return redirect(url_for('students'))
        except Exception as e:
            flash(f'Error reading CSV: {str(e)}', 'danger')
            return redirect(url_for('upload_csv'))
    return render_template('upload_csv.html', user_name=session['user_name'])


# ── COMPANIES ─────────────────────────────────────────────────────────────────
@app.route('/companies')
@login_required
def companies():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM companies ORDER BY visit_date DESC")
    all_companies = cur.fetchall()
    cur.close()
    return render_template('companies.html', companies=all_companies, user_name=session['user_name'])


@app.route('/add_company', methods=['GET', 'POST'])
@admin_required
def add_company():
    if request.method == 'POST':
        try:
            package = float(request.form['package'])
        except ValueError:
            flash('Package must be a number.', 'danger')
            return render_template('add_company.html', user_name=session['user_name'])
        cur = mysql.connection.cursor()
        cur.execute(
            "INSERT INTO companies(company_name,package,required_skills,visit_date) VALUES(%s,%s,%s,%s)",
            (request.form['company_name'], package,
             request.form['required_skills'], request.form['visit_date'])
        )
        mysql.connection.commit()
        cur.close()
        flash('Company added successfully!', 'success')
        return redirect(url_for('companies'))
    return render_template('add_company.html', user_name=session['user_name'])


@app.route('/edit_company/<int:company_id>', methods=['GET', 'POST'])
@admin_required
def edit_company(company_id):
    cur = mysql.connection.cursor()
    if request.method == 'POST':
        try:
            package = float(request.form['package'])
        except ValueError:
            cur.close()
            flash('Package must be a number.', 'danger')
            return redirect(url_for('edit_company', company_id=company_id))
        cur.execute("SELECT 1 FROM companies WHERE company_id=%s", (company_id,))
        if not cur.fetchone():
            cur.close()
            flash('Company not found — it may have already been deleted.', 'danger')
            return redirect(url_for('companies'))
        cur.execute("""
            UPDATE companies SET company_name=%s, package=%s, required_skills=%s, visit_date=%s
            WHERE company_id=%s
        """, (request.form['company_name'], package,
              request.form['required_skills'], request.form['visit_date'], company_id))
        mysql.connection.commit()
        cur.close()
        flash('Company updated successfully!', 'success')
        return redirect(url_for('companies'))
    cur.execute("SELECT * FROM companies WHERE company_id=%s", (company_id,))
    company = cur.fetchone()
    cur.close()
    if not company:
        flash('Company not found.', 'danger')
        return redirect(url_for('companies'))
    return render_template('edit_company.html', company=company, user_name=session['user_name'])


@app.route('/delete_company/<int:company_id>', methods=['POST'])
@admin_required
def delete_company(company_id):
    cur = mysql.connection.cursor()
    try:
        cur.execute("DELETE FROM companies WHERE company_id=%s", (company_id,))
        mysql.connection.commit()
        flash('Company deleted successfully!', 'success')
    except pymysql.err.IntegrityError:
        mysql.connection.rollback()
        flash('Cannot delete this company — it has placement records. Delete those placements first.', 'danger')
    finally:
        cur.close()
    return redirect(url_for('companies'))


# ── PLACEMENTS ────────────────────────────────────────────────────────────────
@app.route('/placements')
@login_required
def placements():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT p.placement_id, s.name, c.company_name, c.package, p.year, p.status
        FROM placements p
        JOIN students s ON p.student_id=s.student_id
        JOIN companies c ON p.company_id=c.company_id
        ORDER BY p.year DESC
    """)
    all_placements = cur.fetchall()
    cur.close()
    return render_template('placements.html', placements=all_placements, user_name=session['user_name'])


@app.route('/add_placement', methods=['GET', 'POST'])
@admin_required
def add_placement():
    cur = mysql.connection.cursor()
    if request.method == 'POST':
        try:
            student_id = int(request.form['student_id'])
            company_id = int(request.form['company_id'])
            year       = int(request.form['year'])
        except ValueError:
            cur.close()
            flash('Please select a valid student, company, and year.', 'danger')
            return redirect(url_for('add_placement'))
        status = request.form['status']
        try:
            cur.execute(
                "INSERT INTO placements(student_id,company_id,year,status) VALUES(%s,%s,%s,%s)",
                (student_id, company_id, year, status)
            )
            mysql.connection.commit()
        except pymysql.err.IntegrityError:
            mysql.connection.rollback()
            cur.close()
            flash('That student or company no longer exists.', 'danger')
            return redirect(url_for('add_placement'))
        cur.execute("SELECT name, email FROM students WHERE student_id=%s", (student_id,))
        student = cur.fetchone()
        cur.execute("SELECT company_name, package FROM companies WHERE company_id=%s", (company_id,))
        company = cur.fetchone()
        cur.close()
        if student and company and os.environ.get('MAIL_USERNAME'):
            try:
                msg = Message('🎉 Congratulations! Placement Confirmed',
                    sender=os.environ.get('MAIL_USERNAME'), recipients=[student[1]])
                msg.body = f'''Dear {student[0]},

🎉 You have been placed at {company[0]}!

Company : {company[0]}
Package : {company[1]} LPA
Year    : {year}
Status  : {status}

Best Regards, Placement Analytics Team'''
                mail.send(msg)
            except Exception:
                pass
        flash('Placement recorded successfully!', 'success')
        return redirect(url_for('placements'))
    cur.execute("SELECT student_id, name FROM students ORDER BY name")
    students = cur.fetchall()
    cur.execute("SELECT company_id, company_name FROM companies ORDER BY company_name")
    companies = cur.fetchall()
    cur.close()
    return render_template('add_placement.html',
        students=students, companies=companies, user_name=session['user_name'])


@app.route('/delete_placement/<int:placement_id>', methods=['POST'])
@admin_required
def delete_placement(placement_id):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM placements WHERE placement_id=%s", (placement_id,))
    mysql.connection.commit()
    cur.close()
    flash('Placement deleted successfully!', 'success')
    return redirect(url_for('placements'))


# ── ANALYTICS ─────────────────────────────────────────────────────────────────
@app.route('/analytics')
@admin_required
def analytics():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT c.company_name, COUNT(*) AS cnt
        FROM placements p JOIN companies c ON p.company_id=c.company_id
        GROUP BY c.company_name ORDER BY cnt DESC LIMIT 10
    """)
    cd = cur.fetchall()
    cur.execute("SELECT year, COUNT(*) AS cnt FROM placements GROUP BY year ORDER BY year")
    yd = cur.fetchall()
    cur.execute("""
        SELECT s.branch, COUNT(*) AS cnt
        FROM placements p JOIN students s ON p.student_id=s.student_id
        GROUP BY s.branch ORDER BY cnt DESC
    """)
    bd = cur.fetchall()
    cur.execute("""
        SELECT c.company_name, c.package
        FROM placements p JOIN companies c ON p.company_id=c.company_id
        ORDER BY c.package DESC LIMIT 10
    """)
    pd = cur.fetchall()
    cur.execute("SELECT required_skills FROM companies")
    all_skills = []
    for row in cur.fetchall():
        if row[0]: all_skills.extend([s.strip() for s in row[0].split(',')])
    sc = Counter(all_skills).most_common(8)
    cur.close()

    company_labels = [r[0] for r in cd]
    company_counts = [r[1] for r in cd]
    year_labels    = [str(r[0]) for r in yd]
    year_counts    = [r[1] for r in yd]
    branch_labels  = [r[0] for r in bd]
    branch_counts  = [r[1] for r in bd]
    pkg_labels     = [r[0] for r in pd]
    pkg_values     = [r[1] for r in pd]
    skill_labels   = [s[0] for s in sc]
    skill_values   = [s[1] for s in sc]

    return render_template('analytics.html',
        company_labels=company_labels,
        company_counts=company_counts,
        year_labels=year_labels,
        year_counts=year_counts,
        branch_labels=branch_labels,
        branch_counts=branch_counts,
        pkg_labels=pkg_labels,
        pkg_values=pkg_values,
        skill_labels=skill_labels,
        skill_values=skill_values,
        kpi_companies=len(company_labels),
        kpi_placed=sum(year_counts),
        kpi_branches=len(branch_labels),
        kpi_skills=len(skill_labels),
        user_name=session['user_name'])


# ── ML PREDICTOR ──────────────────────────────────────────────────────────────
@app.route('/predict', methods=['GET', 'POST'])
@login_required
def predict():
    result = None
    if request.method == 'POST':
        try:
            cgpa       = float(request.form['cgpa'])
            skills     = request.form['skills']
            branch     = request.form.get('branch', 'CSE')
            backlogs   = int(request.form.get('backlogs', 0))
            internship = request.form.get('internship', 'no')
            projects   = int(request.form.get('projects', 0))

            skill_list     = [s.strip().lower() for s in skills.split(',') if s.strip()]
            skill_count    = len(skill_list)
            high_demand    = ['python','java','react','node','sql','mysql','javascript',
                              'dsa','c++','machine learning','ml','django','flask',
                              'spring','aws','docker','git']
            matched_skills = sum(1 for s in skill_list if any(h in s for h in high_demand))

            if cgpa >= 9.5:   cgpa_score = 100
            elif cgpa >= 9.0: cgpa_score = 95
            elif cgpa >= 8.5: cgpa_score = 88
            elif cgpa >= 8.0: cgpa_score = 80
            elif cgpa >= 7.5: cgpa_score = 70
            elif cgpa >= 7.0: cgpa_score = 58
            elif cgpa >= 6.5: cgpa_score = 45
            elif cgpa >= 6.0: cgpa_score = 32
            else:             cgpa_score = 15

            skill_score  = min(100, (skill_count * 12) + (matched_skills * 8))
            branch_score = {'CSE':95,'IT':88,'ECE':75,'EEE':65,'Mechanical':55,'Civil':45}.get(branch, 60)

            if backlogs == 0:   backlog_score = 100
            elif backlogs == 1: backlog_score = 70
            elif backlogs == 2: backlog_score = 45
            elif backlogs <= 4: backlog_score = 20
            else:               backlog_score = 5

            internship_score = 100 if internship == 'yes' else 30
            project_score    = min(100, projects * 25)

            chance = round(
                cgpa_score * 0.30 + skill_score * 0.25 + branch_score * 0.15 +
                backlog_score * 0.15 + internship_score * 0.10 + project_score * 0.05
            )

            if chance >= 85:
                company_matches = [
                    {'name':'Google / Amazon','icon':'🚀','color':'#10b981'},
                    {'name':'Microsoft / Adobe','icon':'⭐','color':'#2563eb'},
                    {'name':'TCS / Infosys','icon':'✅','color':'#10b981'},
                ]
            elif chance >= 70:
                company_matches = [
                    {'name':'Wipro / HCL','icon':'✅','color':'#10b981'},
                    {'name':'Capgemini / Accenture','icon':'⚡','color':'#f59e0b'},
                    {'name':'TCS / Infosys','icon':'✅','color':'#10b981'},
                ]
            elif chance >= 50:
                company_matches = [
                    {'name':'TCS / Infosys','icon':'✅','color':'#10b981'},
                    {'name':'Wipro / Tech Mahindra','icon':'⚡','color':'#f59e0b'},
                ]
            else:
                company_matches = [{'name':'Focus on improving skills first','icon':'📚','color':'#ef4444'}]

            tips = []
            if cgpa_score < 70:    tips.append('📈 Improve your CGPA — target 7.5+')
            if skill_count < 4:    tips.append('💻 Learn more in-demand skills (Python, DSA, SQL)')
            if matched_skills < 3: tips.append('🎯 Focus on: Python, Java, React, DSA')
            if backlogs > 0:       tips.append('📋 Clear all backlogs — they reduce chances significantly')
            if internship == 'no': tips.append('🏢 Try to get an internship — boosts chances by 10%')
            if projects < 2:       tips.append('🚀 Build at least 2-3 real projects')
            if not tips:           tips.append('🌟 You are well prepared — keep practicing DSA!')

            if chance >= 80:   level, color, emoji = 'High',     'green',  '🔥'
            elif chance >= 60: level, color, emoji = 'Medium',   'orange', '⚡'
            elif chance >= 40: level, color, emoji = 'Low',      'red',    '📚'
            else:              level, color, emoji = 'Very Low', 'red',    '😟'

            result = {
                'chance':chance, 'level':level, 'color':color, 'emoji':emoji,
                'cgpa':cgpa, 'skills':skill_count, 'matched_skills':matched_skills,
                'branch':branch, 'backlogs':backlogs, 'internship':internship,
                'projects':projects,
                'scores':{'cgpa':round(cgpa_score),'skills':round(skill_score),
                          'branch':round(branch_score),'backlog':round(backlog_score),
                          'internship':round(internship_score),'projects':round(project_score)},
                'company_matches':company_matches, 'tips':tips
            }
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
    return render_template('predict.html', result=result, user_name=session['user_name'])


# ── PDF REPORT ────────────────────────────────────────────────────────────────
@app.route('/download_report')
@admin_required
def download_report():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT s.name, c.company_name, c.package, p.year, p.status, s.branch
        FROM placements p
        JOIN students s ON p.student_id=s.student_id
        JOIN companies c ON p.company_id=c.company_id ORDER BY p.year DESC
    """)
    pl = cur.fetchall()
    cur.execute("""
        SELECT (SELECT COUNT(*) FROM students) AS ts,
               (SELECT COUNT(*) FROM placements) AS tp,
               (SELECT AVG(c.package) FROM placements p JOIN companies c ON p.company_id=c.company_id) AS ap,
               (SELECT MAX(c.package) FROM placements p JOIN companies c ON p.company_id=c.company_id) AS mp
    """)
    s = cur.fetchone(); cur.close()
    ts, tp, ap, mp = s[0] or 0, s[1] or 0, s[2] or 0, s[3] or 0

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    els = [Paragraph('Smart Placement Analytics Report', styles['Title']), Spacer(1,20),
           Paragraph('Summary Statistics', styles['Heading2']), Spacer(1,10)]
    sd = [['Metric','Value'],['Total Students',str(ts)],['Students Placed',str(tp)],
          ['Average Package',f'{round(ap,2)} LPA'],['Highest Package',f'{round(mp,2)} LPA'],
          ['Placement Rate',f'{round((tp/ts)*100,1) if ts else 0}%']]
    st = Table(sd, colWidths=[250,250])
    st.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#2563eb')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,0),12),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('GRID',(0,0),(-1,-1),1,colors.grey),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f0f4ff')]),
        ('FONTNAME',(0,1),(-1,-1),'Helvetica'),
        ('FONTSIZE',(0,1),(-1,-1),11),
        ('PADDING',(0,0),(-1,-1),8),
    ]))
    els.extend([st, Spacer(1,30), Paragraph('Placement Records', styles['Heading2']), Spacer(1,10)])
    if pl:
        pd2 = [['Student','Company','Package (LPA)','Year','Branch','Status']]
        for p in pl: pd2.append([str(p[0]),str(p[1]),str(p[2]),str(p[3]),str(p[5]),str(p[4])])
        pt = Table(pd2, colWidths=[110,100,80,50,70,80])
        pt.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1e293b')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,0),10),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('GRID',(0,0),(-1,-1),0.5,colors.grey),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f8fafc')]),
            ('FONTNAME',(0,1),(-1,-1),'Helvetica'),
            ('FONTSIZE',(0,1),(-1,-1),9),
            ('PADDING',(0,0),(-1,-1),6),
        ]))
        els.append(pt)
    else:
        els.append(Paragraph('No placement records found.', styles['Normal']))
    doc.build(els); buf.seek(0)
    resp = make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = 'attachment; filename=placement_report.pdf'
    return resp


# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────
def excel_safe(val):
    """Neutralize formula-injection payloads (CWE-1236) before writing to a cell.

    name/skills/company_name reach here from admin input and bulk CSV upload.
    A value like =cmd|'/c calc'!A1 sits in the DB as plain text but becomes a
    live formula the moment someone opens the exported .xlsx in Excel.
    """
    if isinstance(val, str) and val[:1] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + val
    return val

@app.route('/export_excel')
@admin_required
def export_excel():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT s.name,s.email,s.branch,s.cgpa,s.skills,c.company_name,c.package,p.year,p.status
        FROM placements p JOIN students s ON p.student_id=s.student_id
        JOIN companies c ON p.company_id=c.company_id ORDER BY p.year DESC
    """)
    pls = cur.fetchall()
    cur.execute("SELECT name,email,branch,cgpa,skills FROM students ORDER BY name")
    sts = cur.fetchall()
    cur.execute("SELECT company_name,package,required_skills,visit_date FROM companies ORDER BY company_name")
    cos = cur.fetchall()
    cur.close()
    wb = openpyxl.Workbook()
    center = Alignment(horizontal='center', vertical='center')
    def mh(ws, headers, color):
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        font = Font(color='FFFFFF', bold=True, size=11)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = fill; c.font = font; c.alignment = center
    ws1 = wb.active; ws1.title = 'Placements'
    h1 = ['Student','Email','Branch','CGPA','Skills','Company','Package (LPA)','Year','Status']
    mh(ws1, h1, '2563EB')
    for ri, row in enumerate(pls, 2):
        for ci, val in enumerate(row, 1):
            cell = ws1.cell(row=ri, column=ci, value=excel_safe(val)); cell.alignment = center
            if ri % 2 == 0: cell.fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    for col in range(1, len(h1)+1): ws1.column_dimensions[get_column_letter(col)].width = 18
    ws2 = wb.create_sheet('Students')
    mh(ws2, ['Name','Email','Branch','CGPA','Skills'], '1E293B')
    for ri, row in enumerate(sts, 2):
        for ci, val in enumerate(row, 1): ws2.cell(row=ri, column=ci, value=excel_safe(val)).alignment = center
    for col in range(1, 6): ws2.column_dimensions[get_column_letter(col)].width = 20
    ws3 = wb.create_sheet('Companies')
    mh(ws3, ['Company Name','Package (LPA)','Required Skills','Visit Date'], '10B981')
    for ri, row in enumerate(cos, 2):
        for ci, val in enumerate(row, 1): ws3.cell(row=ri, column=ci, value=excel_safe(str(val) if val else '')).alignment = center
    for col in range(1, 5): ws3.column_dimensions[get_column_letter(col)].width = 22
    out = BytesIO(); wb.save(out); out.seek(0)
    resp = make_response(out.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = 'attachment; filename=placement_data.xlsx'
    return resp


# ── PROFILE ───────────────────────────────────────────────────────────────────
@app.route('/profile')
@login_required
def profile():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM students WHERE student_id=%s", (session['user_id'],))
    student = cur.fetchone()
    cur.execute("""
        SELECT c.company_name, c.package, p.year, p.status
        FROM placements p JOIN companies c ON p.company_id=c.company_id
        WHERE p.student_id=%s ORDER BY p.year DESC
    """, (session['user_id'],))
    my_placements = cur.fetchall()
    cur.close()
    return render_template('profile.html',
        student=student, my_placements=my_placements, user_name=session['user_name'])


# ── API ───────────────────────────────────────────────────────────────────────
@app.route('/api/stats')
@login_required
def api_stats():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT (SELECT COUNT(*) FROM students) AS s,
               (SELECT COUNT(*) FROM placements) AS p,
               (SELECT AVG(c.package) FROM placements pl JOIN companies c ON pl.company_id=c.company_id) AS a
    """)
    r = cur.fetchone(); cur.close()
    students, placed, avg_pkg = r[0] or 0, r[1] or 0, r[2] or 0
    return jsonify({
        'total_students': students, 'total_placed': placed,
        'placement_rate': round((placed/students)*100,1) if students else 0,
        'avg_package': round(avg_pkg, 2)
    })


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)