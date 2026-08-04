from io import BytesIO

import openpyxl
from flask import make_response
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from ..extensions import mysql
from ..decorators import admin_required
from ..utils import excel_safe


def register(app):
    @app.route('/download_report')
    @admin_required
    def download_report():
        cur = mysql.connection.cursor()
        cur.execute("""
            SELECT s.name, c.company_name, c.package, p.year, p.status, s.branch
            FROM placements p
            JOIN students s ON p.student_id=s.student_id
            JOIN companies c ON p.company_id=c.company_id ORDER BY p.year DESC
        """)
        pl = cur.fetchall()
        cur.execute("""
            SELECT (SELECT COUNT(*) FROM students) AS ts,
                   (SELECT COUNT(*) FROM placements) AS tp,
                   (SELECT AVG(c.package) FROM placements p JOIN companies c ON p.company_id=c.company_id) AS ap,
                   (SELECT MAX(c.package) FROM placements p JOIN companies c ON p.company_id=c.company_id) AS mp
        """)
        s = cur.fetchone(); cur.close()
        ts, tp, ap, mp = s[0] or 0, s[1] or 0, s[2] or 0, s[3] or 0

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter)
        styles = getSampleStyleSheet()
        els = [Paragraph('Smart Placement Analytics Report', styles['Title']), Spacer(1,20),
               Paragraph('Summary Statistics', styles['Heading2']), Spacer(1,10)]
        sd = [['Metric','Value'],['Total Students',str(ts)],['Students Placed',str(tp)],
              ['Average Package',f'{round(ap,2)} LPA'],['Highest Package',f'{round(mp,2)} LPA'],
              ['Placement Rate',f'{round((tp/ts)*100,1) if ts else 0}%']]
        st = Table(sd, colWidths=[250,250])
        st.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#2563eb')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,0),12),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('GRID',(0,0),(-1,-1),1,colors.grey),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f0f4ff')]),
            ('FONTNAME',(0,1),(-1,-1),'Helvetica'),
            ('FONTSIZE',(0,1),(-1,-1),11),
            ('PADDING',(0,0),(-1,-1),8),
        ]))
        els.extend([st, Spacer(1,30), Paragraph('Placement Records', styles['Heading2']), Spacer(1,10)])
        if pl:
            pd2 = [['Student','Company','Package (LPA)','Year','Branch','Status']]
            for p in pl: pd2.append([str(p[0]),str(p[1]),str(p[2]),str(p[3]),str(p[5]),str(p[4])])
            pt = Table(pd2, colWidths=[110,100,80,50,70,80])
            pt.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1e293b')),
                ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('FONTSIZE',(0,0),(-1,0),10),
                ('ALIGN',(0,0),(-1,-1),'CENTER'),
                ('GRID',(0,0),(-1,-1),0.5,colors.grey),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f8fafc')]),
                ('FONTNAME',(0,1),(-1,-1),'Helvetica'),
                ('FONTSIZE',(0,1),(-1,-1),9),
                ('PADDING',(0,0),(-1,-1),6),
            ]))
            els.append(pt)
        else:
            els.append(Paragraph('No placement records found.', styles['Normal']))
        doc.build(els); buf.seek(0)
        resp = make_response(buf.getvalue())
        resp.headers['Content-Type'] = 'application/pdf'
        resp.headers['Content-Disposition'] = 'attachment; filename=placement_report.pdf'
        return resp

    @app.route('/export_excel')
    @admin_required
    def export_excel():
        cur = mysql.connection.cursor()
        cur.execute("""
            SELECT s.name,s.email,s.branch,s.cgpa,s.skills,c.company_name,c.package,p.year,p.status
            FROM placements p JOIN students s ON p.student_id=s.student_id
            JOIN companies c ON p.company_id=c.company_id ORDER BY p.year DESC
        """)
        pls = cur.fetchall()
        cur.execute("SELECT name,email,branch,cgpa,skills FROM students ORDER BY name")
        sts = cur.fetchall()
        cur.execute("SELECT company_name,package,required_skills,visit_date FROM companies ORDER BY company_name")
        cos = cur.fetchall()
        cur.close()
        wb = openpyxl.Workbook()
        center = Alignment(horizontal='center', vertical='center')
        def mh(ws, headers, color):
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            font = Font(color='FFFFFF', bold=True, size=11)
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.fill = fill; c.font = font; c.alignment = center
        ws1 = wb.active; ws1.title = 'Placements'
        h1 = ['Student','Email','Branch','CGPA','Skills','Company','Package (LPA)','Year','Status']
        mh(ws1, h1, '2563EB')
        for ri, row in enumerate(pls, 2):
            for ci, val in enumerate(row, 1):
                cell = ws1.cell(row=ri, column=ci, value=excel_safe(val)); cell.alignment = center
                if ri % 2 == 0: cell.fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
        for col in range(1, len(h1)+1): ws1.column_dimensions[get_column_letter(col)].width = 18
        ws2 = wb.create_sheet('Students')
        mh(ws2, ['Name','Email','Branch','CGPA','Skills'], '1E293B')
        for ri, row in enumerate(sts, 2):
            for ci, val in enumerate(row, 1): ws2.cell(row=ri, column=ci, value=excel_safe(val)).alignment = center
        for col in range(1, 6): ws2.column_dimensions[get_column_letter(col)].width = 20
        ws3 = wb.create_sheet('Companies')
        mh(ws3, ['Company Name','Package (LPA)','Required Skills','Visit Date'], '10B981')
        for ri, row in enumerate(cos, 2):
            for ci, val in enumerate(row, 1): ws3.cell(row=ri, column=ci, value=excel_safe(str(val) if val else '')).alignment = center
        for col in range(1, 5): ws3.column_dimensions[get_column_letter(col)].width = 22
        out = BytesIO(); wb.save(out); out.seek(0)
        resp = make_response(out.getvalue())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = 'attachment; filename=placement_data.xlsx'
        return resp
